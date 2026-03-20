"""
Generate 2026 Opportunities Excel — from OCR pipeline results + costing sheets.
Matches the format of "opportunities created 2025.xlsx" exactly.

Columns:
  A: Year (2026)
  B: Opp. No. (sequential from folder name)
  C: Folder No. (XX-26)
  D: Folder Name (from project folder)
  E: SFDC Opportunity title (generated)
  F: End User (client name)
  G: Latest Comment (blank — user fills)
  H: E+H Reference No (from costing if found)
  I: Value in BHD (from costing sheet)
  J: Status (default: "Follow-up / Evaluation")
  K: Reason for Loss (blank)
  L: Quote Date (from costing sheet)
  M: Order date (blank)
  N: Approx Delivery Date (blank)
  O: Payment Terms (from costing sheet)
  P: Owner (blank — user fills)
  Q: Abie Comments (blank — user fills)
"""

import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ─── Extract data from costing XLSX files ──────────────────────────

def extract_costing_data(filepath: str) -> dict:
    """
    Read a PH Trading costing XLSX and extract:
    - Total BHD value
    - Quote date
    - Customer name / ID
    - Payment terms
    - E+H reference
    - Quote number
    """
    result = {
        "total_bhd": None,
        "quote_date": None,
        "customer_name": None,
        "customer_id": None,
        "payment_terms": None,
        "eh_reference": None,
        "quote_number": None,
    }

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception:
        # Try xlrd for .xls
        try:
            import xlrd
            wb_xls = xlrd.open_workbook(filepath)
            ws = wb_xls.sheet_by_index(0)
            # Basic extraction from xls
            for row in range(min(ws.nrows, 50)):
                for col in range(min(ws.ncols, 20)):
                    val = ws.cell_value(row, col)
                    if isinstance(val, str):
                        if "date" in val.lower() and col + 1 < ws.ncols:
                            next_val = ws.cell_value(row, col + 1)
                            if next_val and not result["quote_date"]:
                                result["quote_date"] = str(next_val)
            return result
        except Exception:
            return result

    for ws_name in wb.sheetnames[:3]:
        ws = wb[ws_name]
        # Scan the sheet for key fields
        for row in ws.iter_rows(min_row=1, max_row=60, max_col=20, values_only=False):
            for cell in row:
                if cell.value is None:
                    continue
                val = str(cell.value).strip()
                val_lower = val.lower()

                # Quote date
                if "date:" in val_lower or val_lower == "date":
                    # Check next cell(s) for date value
                    next_cells = _get_next_cells(ws, cell.row, cell.column, 3)
                    for nc in next_cells:
                        if nc and _looks_like_date(nc):
                            result["quote_date"] = _parse_date(nc)
                            break

                # Customer name
                if val_lower in ("to,", "to"):
                    next_row_cells = _get_row_values(ws, cell.row + 1, 5)
                    for nc in next_row_cells:
                        if nc and len(str(nc)) > 3 and str(nc) not in ("#N/A", "Not Found", "Salutation Choose Contact Person"):
                            result["customer_name"] = str(nc).strip()
                            break

                # Customer ID
                if "customer id" in val_lower:
                    next_cells = _get_next_cells(ws, cell.row, cell.column, 3)
                    for nc in next_cells:
                        if nc and str(nc).strip() not in ("#N/A", "", "None"):
                            result["customer_id"] = str(nc).strip()
                            break

                # Payment terms
                if "payment" in val_lower and "term" in val_lower:
                    next_cells = _get_next_cells(ws, cell.row, cell.column, 5)
                    for nc in next_cells:
                        if nc and len(str(nc)) > 3:
                            result["payment_terms"] = str(nc).strip()
                            break

                # Quote number
                if "quote #" in val_lower or "quote no" in val_lower:
                    next_cells = _get_next_cells(ws, cell.row, cell.column, 3)
                    for nc in next_cells:
                        if nc and str(nc).strip() not in ("0", "", "None", "#N/A"):
                            result["quote_number"] = str(nc).strip()
                            break

                # E+H reference (long numeric strings)
                if isinstance(cell.value, (int, float)) and cell.value > 1e9:
                    if not result["eh_reference"]:
                        result["eh_reference"] = str(int(cell.value))

        # Look for Grand Total / Total BHD in the sheet
        for row in ws.iter_rows(min_row=1, max_row=100, max_col=20, values_only=False):
            for cell in row:
                if cell.value is None:
                    continue
                val = str(cell.value).strip().lower()
                if any(kw in val for kw in ("grand total", "total amount", "total bhd",
                                            "total incl", "net total", "total value")):
                    # Look for numeric value in nearby cells
                    next_cells = _get_next_cells(ws, cell.row, cell.column, 8)
                    for nc in next_cells:
                        if nc and _looks_like_amount(nc):
                            amount = _parse_amount(nc)
                            if amount and amount > 10:  # Minimum sanity check
                                if result["total_bhd"] is None or amount > result["total_bhd"]:
                                    result["total_bhd"] = amount
                                break
                    # Also check same row for amounts
                    for c2 in row:
                        if c2.value and c2.column != cell.column:
                            if _looks_like_amount(c2.value):
                                amount = _parse_amount(c2.value)
                                if amount and amount > 10:
                                    if result["total_bhd"] is None or amount > result["total_bhd"]:
                                        result["total_bhd"] = amount

    wb.close()
    return result


def _get_next_cells(ws, row, col, count):
    """Get values of next N cells to the right."""
    values = []
    for i in range(1, count + 1):
        try:
            val = ws.cell(row, col + i).value
            if val is not None:
                values.append(val)
        except Exception:
            pass
    return values


def _get_row_values(ws, row, count):
    """Get first N non-empty values from a row."""
    values = []
    for col in range(1, count + 10):
        try:
            val = ws.cell(row, col).value
            if val is not None:
                values.append(val)
        except Exception:
            pass
        if len(values) >= count:
            break
    return values


def _looks_like_date(val) -> bool:
    if isinstance(val, datetime):
        return True
    s = str(val)
    return bool(re.search(r'\d{4}-\d{2}-\d{2}', s) or re.search(r'\d{2}/\d{2}/\d{4}', s))


def _parse_date(val):
    if isinstance(val, datetime):
        return val
    s = str(val)
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s.split(".")[0].strip(), fmt)
        except ValueError:
            continue
    return None


def _looks_like_amount(val) -> bool:
    if isinstance(val, (int, float)):
        return True
    s = str(val).replace(",", "").strip()
    return bool(re.match(r'^\d+\.?\d*$', s))


def _parse_amount(val) -> float:
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


# ─── Parse project folder name ─────────────────────────────────────

def parse_project(project_id: str) -> dict:
    """Parse 'EH-15-26 ALBA FIT' into structured fields."""
    result = {"opp_no": None, "folder_name": project_id, "client": "", "type_suffix": ""}

    # Try to extract number: EH-XX-26, EH XX-26, EHS-XX-26, SA-XX-26, etc.
    m = re.match(r'(EH[S]?|SA|OTH)[\s-]*(\d+)[\s-]*(?:26|2026)', project_id)
    if m:
        prefix = m.group(1).upper()
        num = int(m.group(2))
        # SA and OTH projects get offset to avoid collisions with EH numbers
        if prefix.startswith("SA"):
            num += 200  # SA-01-26 → opp 201
        elif prefix.startswith("OTH"):
            num += 300  # OTH-01-26 → opp 301
        result["opp_no"] = num

    # Extract client name (everything after the number prefix)
    parts = re.split(r'(?:EH[S]?|SA|OTH)[\s-]*\d+[\s-]*(?:26|2026)\s*', project_id)
    if len(parts) > 1 and parts[1]:
        name = parts[1].strip()
        # Remove common type suffixes
        for suf in ["FIT", "TIT", "LIT", "TW", "AIT", "DPIT", "PP", "BTU",
                    "CEMS", "SP", "EHS", "EH", "PH", "CL2", "TOC", "DSP", "FEED"]:
            if name.endswith(f" {suf}"):
                result["type_suffix"] = suf
                name = name[: -(len(suf) + 1)].strip()
        result["client"] = name.strip()

    return result


# ─── Known client name cleanup ──────────────────────────────────────

CLIENT_CLEANUP = {
    "ALBA": "Aluminium Bahrain (ALBA)",
    "GPIC": "Gulf Petrochemical Industries Co (GPIC)",
    "BAPCO": "Bapco Refining",
    "BAPCO REFINING": "Bapco Refining",
    "BAPCO UPSTREAM": "Bapco Upstream",
    "BAPCO Airfueling": "Bapco Airfueling",
    "EWA": "Electricity & Water Authority (EWA)",
    "HAMALA": "EWA - Hamala",
    "HAMALA-EH": "EWA - Hamala",
    "VEOLIA": "Veolia Water Technologies",
    "VEOLIA ENERGY": "Veolia Energy",
    "TTSJV": "TTSJV (Taqyiat-Tabreed SJV)",
    "TTSJV 5822": "TTSJV (Taqyiat-Tabreed SJV)",
    "ARLA": "Arla Foods Bahrain S.P.C",
    "ARLA-FIT": "Arla Foods Bahrain S.P.C",
    "ARLA P.Cover": "Arla Foods Bahrain S.P.C",
    "INTERCOL": "Intercol",
    "NOMAC": "NOMAC",
    "SEAPEAK": "Seapeak",
    "XENITT": "Xenitt",
    "AQUA TECH": "Aqua Tech",
    "WABAG": "WABAG",
    "SEDRES": "Sedres/Orbitus",
    "SEDRESS": "Sedres/Orbitus",
    "AHS TRADING WLL": "AHS Trading W.L.L",
    "ABRAQYAH": "Abraqyah",
    "VWT PIT": "Veolia Water Technologies",
    "YATEEM": "Yateem Group",
    "AL MOAYYED": "Al Moayyed",
    "TABREED": "Tabreed",
    "ALDUR": "Aldur",
    "SULB": "SULB Bahrain",
    "ZOHAL": "Zohal",
    "ENGIE": "Al Ezzel Engie O&M",
    "SYSCON": "Syscon",
    "PRUDENT VALVE": "Prudent Valve",
    "PROMAG": "Promag (Direct)",
}


PLACEHOLDER_TERMS = {
    "select payment terms", "select delivery terms", "choose payment",
    "payment terms", "select terms", "choose terms",
}


def _clean_payment_terms(terms: str) -> str:
    """Remove placeholder/template text from payment terms."""
    if not terms:
        return ""
    if terms.strip().lower() in PLACEHOLDER_TERMS:
        return ""
    return terms.strip()


def clean_client_name(raw: str) -> str:
    for key, clean in CLIENT_CLEANUP.items():
        if raw.upper().startswith(key.upper()):
            return clean
    return raw


# ─── Main generator ─────────────────────────────────────────────────

def main():
    PIPELINE_JSON = "pipeline_with_engines_v2.json"
    TEMPLATE_XLSX = "../docs/reference_data/opportunities created 2025.xlsx"
    OUTPUT_XLSX = "../docs/reference_data/opportunities_2026_from_ocr_v2.xlsx"

    print("=" * 70)
    print("  GENERATING 2026 OPPORTUNITIES EXCEL")
    print("=" * 70)

    # Load pipeline results
    with open(PIPELINE_JSON, "r", encoding="utf-8") as f:
        pipeline = json.load(f)

    projects = pipeline["projects"]
    files_by_project = {}
    for fi in pipeline["files"]:
        proj = fi.get("project", "")
        if proj not in files_by_project:
            files_by_project[proj] = []
        files_by_project[proj].append(fi)

    # Find costing files per project (actual file paths for deep reading)
    print(f"\n  Scanning {len(projects)} projects for costing data...")
    project_data = []

    for pid, pinfo in sorted(projects.items()):
        parsed = parse_project(pid)
        if parsed["opp_no"] is None:
            continue

        # Find costing XLSX files
        costing_result = {"total_bhd": None, "quote_date": None, "payment_terms": None,
                         "eh_reference": None, "customer_id": None, "quote_number": None,
                         "customer_name": None}

        for fi in files_by_project.get(pid, []):
            fname = fi["filename"].lower()
            if "costing" in fname and fi["extension"] in (".xlsx", ".xls"):
                try:
                    extracted = extract_costing_data(fi["path"])
                    # Merge (first non-None wins)
                    for key in costing_result:
                        if costing_result[key] is None and extracted.get(key) is not None:
                            costing_result[key] = extracted[key]
                except Exception as e:
                    pass

        client_raw = parsed["client"] or pinfo.get("client_name", "")
        client_clean = clean_client_name(client_raw)

        entry = {
            "year": 2026,
            "opp_no": parsed["opp_no"],
            "folder_no": f"{parsed['opp_no']}-26",
            "folder_name": pid,
            "sfdc_title": f"{client_clean} {parsed['type_suffix']}".strip() if parsed["type_suffix"] else client_clean,
            "end_user": client_clean,
            "latest_comment": "",  # User fills
            "eh_reference": costing_result["eh_reference"] or "",
            "value_bhd": costing_result["total_bhd"],
            "status": "Follow-up / Evaluation",
            "reason_loss": "",
            "quote_date": costing_result["quote_date"],
            "order_date": None,
            "delivery_date": None,
            "payment_terms": _clean_payment_terms(costing_result["payment_terms"] or ""),
            "owner": "",  # User fills
            "abie_comments": "",  # User fills
            "file_count": pinfo["file_count"],
            "costing_sheets": pinfo["costing_sheets"],
        }
        project_data.append(entry)

    project_data.sort(key=lambda x: x["opp_no"])
    print(f"  Found {len(project_data)} opportunities to export")

    # Count how many have BHD values
    with_values = sum(1 for e in project_data if e["value_bhd"])
    with_dates = sum(1 for e in project_data if e["quote_date"])
    with_terms = sum(1 for e in project_data if e["payment_terms"])
    print(f"  With BHD values:    {with_values}/{len(project_data)}")
    print(f"  With quote dates:   {with_dates}/{len(project_data)}")
    print(f"  With payment terms: {with_terms}/{len(project_data)}")

    # ─── Create Excel ───────────────────────────────────────
    print(f"\n  Creating Excel file...")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2026"

    # Styles (matching the 2025 file)
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    money_format = '#,##0.00'
    date_format = 'YYYY-MM-DD'

    # Headers
    headers = [
        ("Year", 6), ("Opp. No.", 10), ("Folder No.", 13),
        ("Folder Name", 35), ("SFDC Opportunity title", 45),
        ("End User", 30), ("Latest Comment", 43),
        ("E+H Refference No", 18), ("Value in BHD", 16),
        ("Status", 24), ("Reason for Loss", 20),
        ("Quote Date", 15), ("Order date", 15),
        ("Approx Delivery Date", 16), ("Payment Terms", 16),
        ("Owner", 13), ("Abie Comments", 21),
    ]

    for col, (name, width) in enumerate(headers, 1):
        cell = ws.cell(1, col, name)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # Data rows
    for i, entry in enumerate(project_data):
        row = i + 2
        ws.cell(row, 1, entry["year"]).border = border
        ws.cell(row, 2, entry["opp_no"]).border = border
        ws.cell(row, 3, entry["folder_no"]).border = border
        ws.cell(row, 4, entry["folder_name"]).border = border
        ws.cell(row, 5, entry["sfdc_title"]).border = border
        ws.cell(row, 6, entry["end_user"]).border = border
        ws.cell(row, 7, entry["latest_comment"]).border = border  # Blank
        ws.cell(row, 8, entry["eh_reference"]).border = border

        # Value in BHD
        c_val = ws.cell(row, 9)
        if entry["value_bhd"]:
            c_val.value = entry["value_bhd"]
            c_val.number_format = money_format
        c_val.border = border

        ws.cell(row, 10, entry["status"]).border = border
        ws.cell(row, 11, entry["reason_loss"]).border = border

        # Quote date
        c_date = ws.cell(row, 12)
        if entry["quote_date"]:
            if isinstance(entry["quote_date"], datetime):
                c_date.value = entry["quote_date"]
                c_date.number_format = date_format
            else:
                c_date.value = str(entry["quote_date"])
        c_date.border = border

        ws.cell(row, 13).border = border  # Order date — blank
        ws.cell(row, 14).border = border  # Delivery date — blank
        ws.cell(row, 15, entry["payment_terms"]).border = border
        ws.cell(row, 16, entry["owner"]).border = border  # Blank
        ws.cell(row, 17, entry["abie_comments"]).border = border  # Blank

    # Freeze panes (header row)
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:Q{len(project_data) + 1}"

    # Save
    wb.save(OUTPUT_XLSX)
    print(f"\n  Saved to: {OUTPUT_XLSX}")
    print(f"  Rows: {len(project_data)} opportunities")

    # Summary
    total_bhd = sum(e["value_bhd"] for e in project_data if e["value_bhd"])
    print(f"\n  Total pipeline value found: BHD {total_bhd:,.2f}")
    print(f"  (Note: {len(project_data) - with_values} projects need manual BHD entry)")
    print(f"{'=' * 70}")


if __name__ == "__main__":
    main()
