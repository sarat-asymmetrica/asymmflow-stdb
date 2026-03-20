"""
Build Canonical Dataset for AsymmFlow Reimagined
=================================================
Merges three data sources into one comprehensive, clean view:
  1. Old DB (ph_holdings.db) — customer grades, invoices, orders, supplier data
  2. 2025 Excel (Abhie-cleaned) — 435 opportunities with status/outcome
  3. 2026 OCR extraction — 65 new opportunities from automated pipeline

Outputs:
  - canonical_seed.xlsx — Multi-sheet Excel with all entities
  - canonical_seed.json — Machine-readable for STDB seeding

Entities:
  - Parties (customers + suppliers, unified)
  - Pipeline (opportunities across 2024-2026)
  - Orders (from old DB)
  - Invoices (from old DB)
  - Supplier Invoices (from old DB)
"""

import json
import hashlib
import sqlite3
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DB_PATH = r"C:/Projects/ph-final/ph_holdings/ph_holdings.db"
XL25_PATH = r"../docs/reference_data/opportunities created 2025.xlsx"
XL26_PATH = r"../docs/reference_data/opportunities_2026_from_ocr_v2.xlsx"
OUTPUT_XLSX = r"../docs/reference_data/canonical_seed.xlsx"
OUTPUT_JSON = r"../docs/reference_data/canonical_seed.json"

# ── Styles ──
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL_BLUE = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FILL_GREEN = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
HEADER_FILL_ORANGE = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
HEADER_FILL_RED = PatternFill(start_color="843C0C", end_color="843C0C", fill_type="solid")
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# ── Customer name normalization ──
NAME_ALIASES = {
    "ALBA": "Aluminium Bahrain (ALBA)",
    "ALUMINIUM BAHRAIN": "Aluminium Bahrain (ALBA)",
    "ALUMINIUM BAHRAIN (ALBA)": "Aluminium Bahrain (ALBA)",
    "ALUMINIUM BAHRAIN B.S.C (C)": "Aluminium Bahrain (ALBA)",
    "GPIC": "Gulf Petrochemical Industries Co (GPIC)",
    "GULF PETROCHEMICAL INDUSTRIES": "Gulf Petrochemical Industries Co (GPIC)",
    "GULF PETROCHEMICAL INDUSTRIES CO (GPIC)": "Gulf Petrochemical Industries Co (GPIC)",
    "BAPCO REFINING": "Bapco Refining",
    "BAPCO": "Bapco Refining",
    "BAPCO UPSTREAM": "Bapco Upstream",
    "BAPCO AIRFUELING": "Bapco Airfueling",
    "EWA": "Electricity & Water Authority (EWA)",
    "ELECTRICITY & WATER AUTHORITY": "Electricity & Water Authority (EWA)",
    "ELECTRICITY & WATER AUTHORITY (EWA)": "Electricity & Water Authority (EWA)",
    "VEOLIA": "Veolia Water Technologies",
    "VEOLIA WATER TECHNOLOGIES": "Veolia Water Technologies",
    "VWT": "Veolia Water Technologies",
    "VEOLIA ENERGY": "Veolia Energy",
    "TTSJV": "TTSJV",
    "TAQYIAT-TABREED SJV": "TTSJV",
    "ARLA": "Arla Foods Bahrain S.P.C",
    "ARLA FOODS": "Arla Foods Bahrain S.P.C",
    "ARLA FOODS BAHRAIN": "Arla Foods Bahrain S.P.C",
    "ARLA FOODS BAHRAIN S.P.C": "Arla Foods Bahrain S.P.C",
    "INTERCOL": "Intercol",
    "NOMAC": "NOMAC",
    "SEAPEAK": "Seapeak",
    "MUHARRAQ WASTEWATER": "Muharraq Wastewater Services Co",
    "MUHARRAQ WASTEWATER SERVICES CO": "Muharraq Wastewater Services Co",
    "AL EZZEL": "Al Ezzel Engie O&M",
    "AL EZZEL ENGIE": "Al Ezzel Engie O&M",
    "AL EZZEL ENGIE O&M": "Al Ezzel Engie O&M",
    "TABREED": "Tabreed",
    "MINISTRY OF WORKS": "Ministry of Works",
    "XENITT": "Xenitt",
    "WABAG": "WABAG",
    "SULB": "SULB Bahrain",
    "SULB BAHRAIN": "SULB Bahrain",
    "YATEEM": "Yateem Group",
    "YATEEM GROUP": "Yateem Group",
    "AL MOAYYED": "Al Moayyed Group",
    "ABRAQYAH": "Abraqyah",
    "AQUA TECH": "Aqua Tech",
    "ZOHAL": "Zohal",
    "ALDUR": "Aldur",
    "SEDRES": "Sedres/Orbitus",
    "ORBITUS": "Sedres/Orbitus",
    "SYSCON": "Syscon",
    "AHS TRADING": "AHS Trading W.L.L",
    "AHS TRADING W.L.L": "AHS Trading W.L.L",
    "PRUDENT VALVE": "Prudent Valve",
}


def normalize_customer(name):
    if not name:
        return ""
    clean = str(name).strip()
    upper = clean.upper()
    # Try exact match
    if upper in NAME_ALIASES:
        return NAME_ALIASES[upper]
    # Try prefix match
    for key, val in NAME_ALIASES.items():
        if upper.startswith(key):
            return val
    return clean


def safe_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s.split(".")[0].strip(), fmt)
        except ValueError:
            continue
    return None


def safe_float(val):
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def write_sheet(ws, headers, data, fill):
    """Write a sheet with headers and data rows."""
    for col, (name, width) in enumerate(headers, 1):
        cell = ws.cell(1, col, name)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row_idx, col_idx, val)
            cell.border = BORDER
            if isinstance(val, float):
                cell.number_format = "#,##0.00"
            elif isinstance(val, datetime):
                cell.number_format = "YYYY-MM-DD"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{len(data) + 1}"


def main():
    print("=" * 75)
    print("  BUILDING CANONICAL DATASET FOR ASYMMFLOW V5")
    print("=" * 75)

    db = sqlite3.connect(DB_PATH)
    cur = db.cursor()

    # ════════════════════════════════════════════════════════════
    # SHEET 1: PARTIES (Unified Customers + Suppliers)
    # ════════════════════════════════════════════════════════════
    print("\n[1/5] Building PARTIES...")

    parties = {}  # normalized_name → party dict

    # From old DB customers
    cur.execute("""
        SELECT business_name, customer_code, customer_type, customer_grade,
               payment_grade, payment_terms_days, city, country, phone, email,
               total_orders_value, total_orders_count, avg_order_value,
               outstanding_bhd, credit_limit_bhd, has_abb_competition
        FROM customers WHERE deleted_at IS NULL
    """)
    for row in cur.fetchall():
        name = normalize_customer(row[0])
        if not name:
            continue
        if name not in parties:
            parties[name] = {
                "name": name, "code": row[1] or "", "type": "Customer",
                "category": row[2] or "", "grade": row[3] or "",
                "payment_grade": row[4] or "", "payment_terms_days": row[5],
                "city": row[6] or "", "country": row[7] or "Bahrain",
                "phone": row[8] or "", "email": row[9] or "",
                "total_orders_bhd": row[10], "total_orders_count": row[11],
                "avg_order_bhd": row[12], "outstanding_bhd": row[13],
                "credit_limit_bhd": row[14], "has_abb_competition": bool(row[15]),
                "source": "old_db", "active_2024": False, "active_2025": False,
                "active_2026": False,
            }

    # From 2025 Excel
    wb25 = openpyxl.load_workbook(XL25_PATH, data_only=True)
    ws25 = wb25.active
    for row in range(2, ws25.max_row + 1):
        name_raw = ws25.cell(row, 6).value
        year = ws25.cell(row, 1).value
        if not name_raw:
            continue
        name = normalize_customer(name_raw)
        if name not in parties:
            parties[name] = {
                "name": name, "code": "", "type": "Customer", "category": "",
                "grade": "", "payment_grade": "", "payment_terms_days": None,
                "city": "", "country": "Bahrain", "phone": "", "email": "",
                "total_orders_bhd": None, "total_orders_count": None,
                "avg_order_bhd": None, "outstanding_bhd": None,
                "credit_limit_bhd": None, "has_abb_competition": False,
                "source": "2025_excel", "active_2024": False,
                "active_2025": False, "active_2026": False,
            }
        if year == 2024:
            parties[name]["active_2024"] = True
        if year == 2025:
            parties[name]["active_2025"] = True

    # From 2026 OCR
    wb26 = openpyxl.load_workbook(XL26_PATH, data_only=True)
    ws26 = wb26.active
    for row in range(2, ws26.max_row + 1):
        name_raw = ws26.cell(row, 6).value
        if not name_raw:
            continue
        name = normalize_customer(name_raw)
        if name not in parties:
            parties[name] = {
                "name": name, "code": "", "type": "Customer", "category": "",
                "grade": "", "payment_grade": "", "payment_terms_days": None,
                "city": "", "country": "Bahrain", "phone": "", "email": "",
                "total_orders_bhd": None, "total_orders_count": None,
                "avg_order_bhd": None, "outstanding_bhd": None,
                "credit_limit_bhd": None, "has_abb_competition": False,
                "source": "2026_ocr", "active_2024": False,
                "active_2025": False, "active_2026": False,
            }
        parties[name]["active_2026"] = True

    # From old DB suppliers
    cur.execute("SELECT id, created_at, deleted_at FROM suppliers WHERE deleted_at IS NULL")
    supplier_count = len(cur.fetchall())

    print(f"  Customers: {len(parties)} (merged from all sources)")
    print(f"  Suppliers in old DB: {supplier_count}")

    active_all = sum(1 for p in parties.values() if p["active_2024"] and p["active_2025"] and p["active_2026"])
    active_26 = sum(1 for p in parties.values() if p["active_2026"])
    with_grade = sum(1 for p in parties.values() if p["grade"])
    print(f"  With grade: {with_grade}, Active 2026: {active_26}, Active all years: {active_all}")

    # ════════════════════════════════════════════════════════════
    # SHEET 2: PIPELINE (All opportunities 2024-2026)
    # ════════════════════════════════════════════════════════════
    print("\n[2/5] Building PIPELINE...")

    pipeline = []

    # From 2025 Excel (2024 + 2025 entries)
    for row in range(2, ws25.max_row + 1):
        year = ws25.cell(row, 1).value
        opp_no = ws25.cell(row, 2).value
        folder = ws25.cell(row, 3).value
        folder_name = ws25.cell(row, 4).value
        sfdc = ws25.cell(row, 5).value
        client_raw = ws25.cell(row, 6).value
        comment = ws25.cell(row, 7).value
        eh_ref = ws25.cell(row, 8).value
        bhd = safe_float(ws25.cell(row, 9).value)
        status = ws25.cell(row, 10).value
        loss_reason = ws25.cell(row, 11).value
        quote_date = safe_date(ws25.cell(row, 12).value)
        order_date = safe_date(ws25.cell(row, 13).value)
        delivery_date = ws25.cell(row, 14).value
        terms = ws25.cell(row, 15).value
        owner = ws25.cell(row, 16).value
        abie = ws25.cell(row, 17).value

        client = normalize_customer(client_raw)
        grade = parties.get(client, {}).get("grade", "")

        pipeline.append([
            year, opp_no, str(folder or ""), str(folder_name or ""),
            str(sfdc or ""), client, grade,
            str(comment or ""), str(eh_ref or ""), bhd,
            str(status or ""), str(loss_reason or ""),
            quote_date, order_date, str(delivery_date or ""),
            str(terms or ""), str(owner or ""), str(abie or ""),
            "2025_excel",
        ])

    # From 2026 OCR
    for row in range(2, ws26.max_row + 1):
        opp_no = ws26.cell(row, 2).value
        folder = ws26.cell(row, 3).value
        folder_name = ws26.cell(row, 4).value
        sfdc = ws26.cell(row, 5).value
        client_raw = ws26.cell(row, 6).value
        eh_ref = ws26.cell(row, 8).value
        bhd = safe_float(ws26.cell(row, 9).value)
        quote_date = safe_date(ws26.cell(row, 12).value)
        terms = ws26.cell(row, 15).value

        client = normalize_customer(client_raw)
        grade = parties.get(client, {}).get("grade", "")

        pipeline.append([
            2026, opp_no, str(folder or ""), str(folder_name or ""),
            str(sfdc or ""), client, grade,
            "", str(eh_ref or ""), bhd,
            "Follow-up / Evaluation", "",
            quote_date, None, "",
            str(terms or ""), "", "",
            "2026_ocr",
        ])

    print(f"  Total pipeline entries: {len(pipeline)}")
    by_year = Counter(r[0] for r in pipeline if r[0])
    for y, c in sorted(by_year.items()):
        print(f"    {y}: {c} entries")
    total_bhd = sum(r[9] for r in pipeline if r[9])
    print(f"  Total pipeline BHD: {total_bhd:,.2f}")

    # ════════════════════════════════════════════════════════════
    # SHEET 3: ORDERS (from old DB)
    # ════════════════════════════════════════════════════════════
    print("\n[3/5] Building ORDERS...")
    cur.execute("""
        SELECT order_number, customer_name, customer_po_number, order_date,
               grand_total_bhd, status, payment_terms, delivery_terms,
               offer_number, delivery_weeks
        FROM orders WHERE deleted_at IS NULL
        ORDER BY order_date DESC
    """)
    orders = []
    for row in cur.fetchall():
        orders.append([
            row[0], normalize_customer(row[1]), row[2],
            safe_date(row[3]), safe_float(row[4]),
            row[5], row[6] or "", row[7] or "",
            row[8] or "", row[9] or "",
        ])
    print(f"  Orders: {len(orders)}")

    # ════════════════════════════════════════════════════════════
    # SHEET 4: INVOICES (from old DB)
    # ════════════════════════════════════════════════════════════
    print("\n[4/5] Building INVOICES...")
    cur.execute("""
        SELECT invoice_number, customer_name, invoice_date, due_date,
               grand_total_bhd, outstanding_bhd, status, payment_terms,
               offer_number, customer_po_number, gross_margin_percent
        FROM invoices WHERE deleted_at IS NULL
        ORDER BY invoice_date DESC
    """)
    invoices = []
    for row in cur.fetchall():
        invoices.append([
            row[0], normalize_customer(row[1]),
            safe_date(row[2]), safe_date(row[3]),
            safe_float(row[4]), safe_float(row[5]),
            row[6], row[7] or "", row[8] or "",
            row[9] or "", safe_float(row[10]),
        ])
    print(f"  Invoices: {len(invoices)}")

    # ════════════════════════════════════════════════════════════
    # SHEET 5: SUPPLIER INVOICES (from old DB)
    # ════════════════════════════════════════════════════════════
    print("\n[5/5] Building SUPPLIER INVOICES...")
    cur.execute("""
        SELECT si.id, si.created_at, si.deleted_at
        FROM supplier_invoices si WHERE si.deleted_at IS NULL
        LIMIT 1
    """)
    # Get columns dynamically
    cur.execute("PRAGMA table_info(supplier_invoices)")
    si_cols = [c[1] for c in cur.fetchall()]

    cur.execute("""
        SELECT * FROM supplier_invoices WHERE deleted_at IS NULL
        ORDER BY created_at DESC LIMIT 500
    """)
    si_rows = cur.fetchall()
    # Find useful columns
    si_useful = []
    for row in si_rows:
        d = dict(zip(si_cols, row))
        si_useful.append([
            d.get("id", "")[:8] if d.get("id") else "",
            d.get("created_at", ""),
            safe_float(d.get("total_amount_bhd")) or safe_float(d.get("grand_total_bhd")),
            d.get("status", ""),
        ])
    print(f"  Supplier invoices: {len(si_useful)}")

    db.close()
    wb25.close()
    wb26.close()

    # ════════════════════════════════════════════════════════════
    # WRITE EXCEL
    # ════════════════════════════════════════════════════════════
    print(f"\nWriting canonical Excel...")
    wb = openpyxl.Workbook()

    # -- Parties sheet --
    ws_parties = wb.active
    ws_parties.title = "Parties"
    party_headers = [
        ("Party Name", 35), ("Code", 10), ("Type", 12), ("Category", 15),
        ("Grade", 7), ("Pay Grade", 10), ("Pay Terms Days", 14),
        ("City", 12), ("Country", 10),
        ("Total Orders BHD", 16), ("Outstanding BHD", 14),
        ("ABB Competes", 11),
        ("Active 2024", 10), ("Active 2025", 10), ("Active 2026", 10),
        ("Source", 12),
    ]
    party_data = []
    for name in sorted(parties.keys()):
        p = parties[name]
        party_data.append([
            p["name"], p["code"], p["type"], p["category"],
            p["grade"], p["payment_grade"], p["payment_terms_days"],
            p["city"], p["country"],
            p["total_orders_bhd"], p["outstanding_bhd"],
            "Yes" if p["has_abb_competition"] else "",
            "Y" if p["active_2024"] else "", "Y" if p["active_2025"] else "",
            "Y" if p["active_2026"] else "",
            p["source"],
        ])
    write_sheet(ws_parties, party_headers, party_data, HEADER_FILL_BLUE)
    print(f"  Parties sheet: {len(party_data)} rows")

    # -- Pipeline sheet --
    ws_pipeline = wb.create_sheet("Pipeline")
    pipeline_headers = [
        ("Year", 6), ("Opp No", 8), ("Folder No", 11), ("Folder Name", 32),
        ("SFDC Title", 40), ("Client", 30), ("Grade", 7),
        ("Comment", 35), ("E+H Ref", 16), ("Value BHD", 14),
        ("Status", 24), ("Loss Reason", 22),
        ("Quote Date", 12), ("Order Date", 12), ("Delivery", 12),
        ("Payment Terms", 18), ("Owner", 10), ("Abie Notes", 18),
        ("Source", 10),
    ]
    write_sheet(ws_pipeline, pipeline_headers, pipeline, HEADER_FILL_GREEN)
    print(f"  Pipeline sheet: {len(pipeline)} rows")

    # -- Orders sheet --
    ws_orders = wb.create_sheet("Orders")
    order_headers = [
        ("Order No", 14), ("Customer", 30), ("Customer PO", 16),
        ("Order Date", 12), ("Total BHD", 14), ("Status", 14),
        ("Payment Terms", 20), ("Delivery Terms", 20),
        ("Offer No", 14), ("Delivery Weeks", 12),
    ]
    write_sheet(ws_orders, order_headers, orders, HEADER_FILL_ORANGE)
    print(f"  Orders sheet: {len(orders)} rows")

    # -- Invoices sheet --
    ws_inv = wb.create_sheet("Invoices")
    inv_headers = [
        ("Invoice No", 14), ("Customer", 30),
        ("Invoice Date", 12), ("Due Date", 12),
        ("Total BHD", 14), ("Outstanding BHD", 14),
        ("Status", 12), ("Payment Terms", 20),
        ("Offer No", 14), ("Customer PO", 14), ("Margin %", 10),
    ]
    write_sheet(ws_inv, inv_headers, invoices, HEADER_FILL_RED)
    print(f"  Invoices sheet: {len(invoices)} rows")

    # -- Summary sheet --
    ws_sum = wb.create_sheet("Summary")
    ws_sum.sheet_properties.tabColor = "4472C4"

    summary_lines = [
        ["CANONICAL SEED DATA — AsymmFlow V5"],
        [f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"],
        [""],
        ["DATA SOURCES:"],
        [f"  1. ph_holdings.db (old V4 snapshot)"],
        [f"  2. opportunities created 2025.xlsx (Abhie-cleaned)"],
        [f"  3. opportunities_2026_from_ocr_v2.xlsx (OCR pipeline, $0.38)"],
        [""],
        ["ENTITY COUNTS:"],
        [f"  Parties (customers): {len(party_data)}"],
        [f"  Pipeline entries: {len(pipeline)} (2024: {by_year.get(2024,0)}, 2025: {by_year.get(2025,0)}, 2026: {by_year.get(2026,0)})"],
        [f"  Orders: {len(orders)}"],
        [f"  Invoices: {len(invoices)}"],
        [""],
        ["FINANCIAL SUMMARY:"],
        [f"  Total pipeline value: BHD {total_bhd:,.2f}"],
        [f"  Total invoiced (history): BHD {sum(r[4] for r in invoices if r[4]):,.2f}"],
        [f"  Total ordered (history): BHD {sum(r[4] for r in orders if r[4]):,.2f}"],
        [""],
        ["CUSTOMER GRADES:"],
        [f"  A (best): {sum(1 for p in parties.values() if p['grade']=='A')}"],
        [f"  B: {sum(1 for p in parties.values() if p['grade']=='B')}"],
        [f"  C: {sum(1 for p in parties.values() if p['grade']=='C')}"],
        [f"  Ungraded: {sum(1 for p in parties.values() if not p['grade'])}"],
        [""],
        ["NOTE: All financial values in BHD (Bahraini Dinar). 10% VAT."],
        ["      Pending Abhie's review for corrections and grade updates."],
    ]
    for i, line in enumerate(summary_lines, 1):
        cell = ws_sum.cell(i, 1, line[0] if line else "")
        if i == 1:
            cell.font = Font(bold=True, size=14)
        elif line and line[0].endswith(":"):
            cell.font = Font(bold=True, size=11)
    ws_sum.column_dimensions["A"].width = 70

    wb.save(OUTPUT_XLSX)
    print(f"\n  Saved Excel: {OUTPUT_XLSX}")

    # ════════════════════════════════════════════════════════════
    # WRITE JSON (for STDB seeding)
    # ════════════════════════════════════════════════════════════
    canonical = {
        "generated": datetime.now().isoformat(),
        "sources": ["ph_holdings.db", "opportunities_2025.xlsx", "ocr_2026_extraction"],
        "parties": [
            {k: (v.isoformat() if isinstance(v, datetime) else v)
             for k, v in p.items()}
            for p in sorted(parties.values(), key=lambda x: x["name"])
        ],
        "pipeline": [
            {
                "year": r[0], "opp_no": r[1], "folder_no": r[2],
                "folder_name": r[3], "sfdc_title": r[4], "client": r[5],
                "grade": r[6], "comment": r[7], "eh_ref": r[8],
                "value_bhd": r[9], "status": r[10], "loss_reason": r[11],
                "quote_date": r[12].isoformat() if isinstance(r[12], datetime) else r[12],
                "order_date": r[13].isoformat() if isinstance(r[13], datetime) else r[13],
                "delivery": r[14], "payment_terms": r[15],
                "owner": r[16], "notes": r[17], "source": r[18],
            }
            for r in pipeline
        ],
        "stats": {
            "parties": len(parties),
            "pipeline_entries": len(pipeline),
            "orders": len(orders),
            "invoices": len(invoices),
            "total_pipeline_bhd": round(total_bhd, 2),
            "by_year": dict(by_year),
        },
    }

    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(canonical, f, indent=2, ensure_ascii=False, default=str)
    print(f"  Saved JSON: {OUTPUT_JSON}")

    # ── Final summary ──
    print(f"\n{'=' * 75}")
    print(f"  CANONICAL DATASET COMPLETE")
    print(f"{'=' * 75}")
    print(f"  Parties:   {len(party_data):>5}")
    print(f"  Pipeline:  {len(pipeline):>5} ({by_year.get(2024,0)} + {by_year.get(2025,0)} + {by_year.get(2026,0)})")
    print(f"  Orders:    {len(orders):>5}")
    print(f"  Invoices:  {len(invoices):>5}")
    print(f"  BHD total: {total_bhd:>12,.2f}")
    print(f"{'=' * 75}")


if __name__ == "__main__":
    main()
