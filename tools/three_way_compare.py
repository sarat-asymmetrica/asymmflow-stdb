"""Three-way comparison: Old DB vs 2025 Excel vs 2026 OCR extraction."""
import sqlite3
import openpyxl
from collections import Counter

DB_PATH = r"C:/Projects/ph-final/ph_holdings/ph_holdings.db"
XL25_PATH = r"../docs/reference_data/opportunities created 2025.xlsx"
XL26_PATH = r"../docs/reference_data/opportunities_2026_from_ocr_v2.xlsx"


def main():
    db = sqlite3.connect(DB_PATH)
    cur = db.cursor()

    print("=" * 75)
    print("  THREE-WAY COMPARISON: Old DB vs 2025 Excel vs 2026 OCR")
    print("=" * 75)

    # ── 1. OLD DB ──
    print("\n--- OLD DATABASE (ph_holdings.db) ---")

    cur.execute("SELECT folder_number, customer_name, revenue_bhd, stage, salesperson FROM opportunities WHERE deleted_at IS NULL")
    db_opps = cur.fetchall()
    print(f"  Opportunities: {len(db_opps)}")

    cur.execute("SELECT offer_number, customer_name, total_value_bhd, stage, payment_terms FROM offers WHERE deleted_at IS NULL")
    db_offers = cur.fetchall()
    print(f"  Offers: {len(db_offers)}")

    cur.execute("SELECT order_number, customer_name, grand_total_bhd, status FROM orders WHERE deleted_at IS NULL")
    db_orders = cur.fetchall()
    print(f"  Orders: {len(db_orders)}")

    cur.execute("SELECT invoice_number, customer_name, grand_total_bhd, status FROM invoices WHERE deleted_at IS NULL")
    db_invoices = cur.fetchall()
    print(f"  Invoices: {len(db_invoices)}")

    cur.execute("SELECT business_name, customer_grade, payment_grade, outstanding_bhd, total_orders_value FROM customers WHERE deleted_at IS NULL")
    db_customers = cur.fetchall()
    print(f"  Customers: {len(db_customers)}")

    cur.execute("SELECT COUNT(*) FROM costing_sheet_data")
    print(f"  Costing sheets: {cur.fetchone()[0]}")
    cur.execute("SELECT COUNT(*) FROM costing_line_items")
    print(f"  Costing line items: {cur.fetchone()[0]}")
    cur.execute("SELECT COUNT(*) FROM supplier_invoices")
    print(f"  Supplier invoices: {cur.fetchone()[0]}")

    total_offered = sum(r[2] for r in db_offers if r[2]) or 0
    total_ordered = sum(r[2] for r in db_orders if r[2]) or 0
    total_invoiced = sum(r[2] for r in db_invoices if r[2]) or 0
    db_opp_total = sum(r[2] for r in db_opps if r[2]) or 0

    print(f"\n  Financial snapshot:")
    print(f"    Opportunities: BHD {db_opp_total:>12,.2f} ({len(db_opps)} deals)")
    print(f"    Offers:        BHD {total_offered:>12,.2f} ({len(db_offers)} offers)")
    print(f"    Orders:        BHD {total_ordered:>12,.2f} ({len(db_orders)} orders)")
    print(f"    Invoices:      BHD {total_invoiced:>12,.2f} ({len(db_invoices)} invoices)")

    db_customer_names = set()
    for c in db_customers:
        if c[0]:
            db_customer_names.add(c[0].strip().upper()[:25])

    # ── 2. 2025 EXCEL ──
    print("\n--- 2025 EXCEL (Abhie-cleaned) ---")
    wb25 = openpyxl.load_workbook(XL25_PATH, data_only=True)
    ws25 = wb25.active

    xl25_customers = set()
    xl25_total = 0
    xl25_count = 0
    xl25_statuses = Counter()
    for row in range(2, ws25.max_row + 1):
        client = ws25.cell(row, 6).value
        bhd = ws25.cell(row, 9).value
        status = ws25.cell(row, 10).value
        if client:
            xl25_customers.add(str(client).strip().upper()[:25])
        if bhd:
            try:
                xl25_total += float(bhd)
                xl25_count += 1
            except (ValueError, TypeError):
                pass
        if status:
            xl25_statuses[str(status).strip()] += 1

    print(f"  Rows: {ws25.max_row - 1}")
    print(f"  Unique customers: {len(xl25_customers)}")
    print(f"  Total pipeline: BHD {xl25_total:>12,.2f} ({xl25_count} with values)")
    print(f"  Statuses: {dict(xl25_statuses.most_common(5))}")

    # ── 3. 2026 OCR ──
    print("\n--- 2026 OCR EXTRACTION (today) ---")
    wb26 = openpyxl.load_workbook(XL26_PATH, data_only=True)
    ws26 = wb26.active

    xl26_customers = set()
    xl26_total = 0
    xl26_count = 0
    for row in range(2, ws26.max_row + 1):
        client = ws26.cell(row, 6).value
        bhd = ws26.cell(row, 9).value
        if client:
            xl26_customers.add(str(client).strip().upper()[:25])
        if bhd:
            try:
                xl26_total += float(bhd)
                xl26_count += 1
            except (ValueError, TypeError):
                pass

    print(f"  Rows: {ws26.max_row - 1}")
    print(f"  Unique customers: {len(xl26_customers)}")
    print(f"  Total pipeline: BHD {xl26_total:>12,.2f} ({xl26_count} with values)")

    # ── 4. CROSS-COMPARISON ──
    print("\n" + "=" * 75)
    print("  CROSS-COMPARISON")
    print("=" * 75)

    all_three = db_customer_names & xl25_customers & xl26_customers
    db_only = db_customer_names - xl25_customers - xl26_customers
    xl25_only = xl25_customers - db_customer_names - xl26_customers
    xl26_only = xl26_customers - db_customer_names - xl25_customers

    print(f"\n  CUSTOMER VENN:")
    print(f"    Old DB:              {len(db_customer_names):>4} customers")
    print(f"    2025 Excel:          {len(xl25_customers):>4} customers")
    print(f"    2026 OCR:            {len(xl26_customers):>4} customers")
    print(f"    In ALL three:        {len(all_three):>4}  <-- core accounts")
    print(f"    Only in old DB:      {len(db_only):>4}")
    print(f"    Only in 2025 Excel:  {len(xl25_only):>4}")
    print(f"    Only in 2026 OCR:    {len(xl26_only):>4}  <-- new business!")

    if all_three:
        print(f"\n  Core accounts (all 3 sources): {sorted(all_three)}")
    if xl26_only:
        print(f"\n  NEW in 2026 (not in DB or 2025): {sorted(xl26_only)[:20]}")

    # Financial comparison
    print(f"\n  PIPELINE VALUE ACROSS SOURCES:")
    print(f"    {'Source':<28} {'Pipeline BHD':>14} {'Deals':>7} {'Avg Deal':>12}")
    print(f"    {'-'*28} {'-'*14} {'-'*7} {'-'*12}")

    sources = [
        ("Old DB (opportunities)", db_opp_total, len(db_opps)),
        ("Old DB (offers)", total_offered, len(db_offers)),
        ("Old DB (orders)", total_ordered, len(db_orders)),
        ("Old DB (invoices)", total_invoiced, len(db_invoices)),
        ("2025 Excel (Abhie)", xl25_total, xl25_count),
        ("2026 OCR (today)", xl26_total, xl26_count),
    ]
    for name, total, count in sources:
        avg = total / max(count, 1)
        print(f"    {name:<28} {total:>13,.2f} {count:>7} {avg:>11,.2f}")

    # Customer grades for 2026 customers
    print(f"\n  2026 CUSTOMERS - GRADE FROM OLD DB:")
    db_cust_lookup = {}
    for row in db_customers:
        if row[0]:
            db_cust_lookup[row[0].strip().upper()[:25]] = {
                "grade": row[1], "pay_grade": row[2],
                "outstanding": row[3], "total_orders": row[4],
            }

    matched = 0
    for c26 in sorted(xl26_customers):
        info = db_cust_lookup.get(c26)
        if info:
            out = f"BHD {info['outstanding']:,.0f}" if info["outstanding"] else "-"
            tot = f"BHD {info['total_orders']:,.0f}" if info["total_orders"] else "-"
            print(f"    {c26:<28} G={str(info['grade'] or '?'):<3} "
                  f"PG={str(info['pay_grade'] or '?'):<3} "
                  f"Outstanding={out:<12} History={tot}")
            matched += 1

    print(f"\n    Matched: {matched}/{len(xl26_customers)} in old DB")
    print(f"    New (no history): {len(xl26_customers) - matched}")

    # Offer stage comparison
    print(f"\n  OFFER STAGES (Old DB):")
    db_stages = Counter(r[3] for r in db_offers if r[3])
    for s, c in db_stages.most_common():
        print(f"    {s:<30} {c:>4}")

    print(f"\n  OPPORTUNITY STAGES (Old DB):")
    db_opp_stages = Counter(r[3] for r in db_opps if r[3])
    for s, c in db_opp_stages.most_common():
        print(f"    {s:<30} {c:>4}")

    # Data quality comparison
    print(f"\n  DATA QUALITY SCORECARD:")
    print(f"    {'Metric':<35} {'Old DB':>10} {'2025 XL':>10} {'2026 OCR':>10}")
    print(f"    {'-'*35} {'-'*10} {'-'*10} {'-'*10}")

    # Customer name fill
    db_cname = sum(1 for r in db_opps if r[1])
    xl25_cname = sum(1 for row in range(2, ws25.max_row+1) if ws25.cell(row, 6).value)
    xl26_cname = sum(1 for row in range(2, ws26.max_row+1) if ws26.cell(row, 6).value)
    print(f"    {'Customer name filled':<35} {db_cname}/{len(db_opps):>5} "
          f"{xl25_cname}/{ws25.max_row-1:>5} {xl26_cname}/{ws26.max_row-1:>5}")

    # Value filled
    db_vfill = sum(1 for r in db_opps if r[2])
    print(f"    {'BHD value filled':<35} {db_vfill}/{len(db_opps):>5} "
          f"{xl25_count}/{ws25.max_row-1:>5} {xl26_count}/{ws26.max_row-1:>5}")

    wb25.close()
    wb26.close()
    db.close()

    print(f"\n{'=' * 75}")
    print(f"  Three-way comparison complete!")
    print(f"{'=' * 75}")


if __name__ == "__main__":
    main()
